# -*- coding: utf-8 -*-

# === 1. Importi ===
from flask import Flask, render_template, request, jsonify, g, Response, send_from_directory, url_for, session, redirect, flash
import datetime
import sqlite3
import os
import uuid
import csv
from io import StringIO, BytesIO
from werkzeug.utils import secure_filename
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from collections import Counter
import base64
import pytz
import functools

 === 2. Konfiguracija ===
app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # folder gde je app.py
DATABASE = os.path.join(BASE_DIR, "evidencija.db")

UPLOAD_FOLDER_GRESKE = os.path.join(app.root_path, 'uploads', 'greske')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER_GRESKE'] = UPLOAD_FOLDER_GRESKE
DELETE_PASSWORD = "Preis0101"

# NOVO: Secret Key je OBAVEZAN za rad sesija. Promijeni ovo u nešto svoje!
app.config['SECRET_KEY'] = 'tvoja-jako-tajna-i-jedinstvena-sifra-123!@#'
app.secret_key = app.config['SECRET_KEY']

if not os.path.exists(app.config['UPLOAD_FOLDER_GRESKE']):
    try:
        os.makedirs(app.config['UPLOAD_FOLDER_GRESKE'])
    except OSError as e:
        print(f"Greška kod kreiranja foldera: {e}")
        
# === NOVO: Definicija korisnika ===
USERS = {
    'icagalj': {
        'password': 'Tech2020',
        'full_name': 'Ivan Cagalj'
    },
    'nglisic': {
        'password': 'PreisNatasa',
        'full_name': 'Natasa Glisic'
    },
    'dbajic': {
        'password': 'Bajic2025',
        'full_name': 'Dario Bajic'
    },
    'aosmic': {
        'password': 'Preis0202',
        'full_name': 'Adin Osmic'
    },
    'sknezevic': {
        'password': 'StipoPreis',
        'full_name': 'Stipo Knezevic'
    },
    'acosic': {
        'password': 'Toni2025PU',
        'full_name': 'Antonio Cosic'
    },
    'abasic': {
        'password': 'BasicAldina2025',
        'full_name': 'Aldina Basic'
    },
    'jurica': {
        'password': 'PU2025',
        'full_name': 'Jurica Djevenica'
    },
    'dzojicm': {
        'password': 'PU2025Mladen',
        'full_name': 'Dzojic Mladen'
    },
}

# === NOVO: Decorator za provjeru prijave ===
def login_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if 'username' not in session:
            flash('Molimo prijavite se za pristup ovoj stranici.', 'warning')
            return redirect(url_for('login'))
        return view(**kwargs)
    return wrapped_view

# === NOVO: Kontekst procesor da su podaci o korisniku dostupni u svim templejtima ===
@app.context_processor
def inject_user():
    if 'username' in session:
        return dict(current_user=USERS.get(session['username']))
    return dict(current_user=None)

# === 3. Baza Podataka ===
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db_exists = os.path.exists(DATABASE)
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        if not db_exists:
            init_db(db)
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db(db=None):
    created_now = False
    if db is None:
        db = sqlite3.connect(DATABASE)
        db.execute("PRAGMA foreign_keys = ON")
        created_now = True
    try:
        cursor = db.cursor()
        try:
            cursor.execute('ALTER TABLE projekti ADD COLUMN kupac TEXT')
        except sqlite3.OperationalError: pass
        cursor.execute('CREATE TABLE IF NOT EXISTS projekti (id INTEGER PRIMARY KEY AUTOINCREMENT, naziv TEXT NOT NULL UNIQUE, kupac TEXT)')
        cursor.execute('CREATE TABLE IF NOT EXISTS sklopovi (id INTEGER PRIMARY KEY AUTOINCREMENT, projekt_id INTEGER NOT NULL, naziv TEXT NOT NULL, FOREIGN KEY (projekt_id) REFERENCES projekti (id) ON DELETE CASCADE)')
        cursor.execute('''CREATE TABLE IF NOT EXISTS greske (id INTEGER PRIMARY KEY AUTOINCREMENT, sklop_id INTEGER NOT NULL, tip TEXT, mjesto TEXT, vrijeme_evidentiranja TEXT NOT NULL, evidentirao TEXT, vrijeme_rjesenja TEXT, rijesio TEXT, FOREIGN KEY (sklop_id) REFERENCES sklopovi (id) ON DELETE CASCADE)''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS greska_slike (id INTEGER PRIMARY KEY AUTOINCREMENT, greska_id INTEGER NOT NULL, naziv_datoteke TEXT NOT NULL, originalni_naziv TEXT, vrijeme_uploada TEXT NOT NULL, FOREIGN KEY (greska_id) REFERENCES greske (id) ON DELETE CASCADE )''')
        db.commit()
    except sqlite3.Error as e:
        print(f"Greška init DB: {e}")
        db.rollback()
    finally:
        if created_now:
            db.close()

# === 3.5 Pomoćne Funkcije ===
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def format_datetime_hr_excel(iso_string, target_timezone='Europe/Zagreb'):
    if not iso_string: return "N/A"
    try:
        dt_object_aware = datetime.datetime.fromisoformat(iso_string.replace('Z', '+00:00'))
        if dt_object_aware.tzinfo is None or dt_object_aware.tzinfo.utcoffset(dt_object_aware) is None:
            dt_object_aware = pytz.utc.localize(dt_object_aware.replace(tzinfo=None))
        target_tz = pytz.timezone(target_timezone)
        dt_object_local = dt_object_aware.astimezone(target_tz)
        return dt_object_local.replace(tzinfo=None)
    except:
        try:
            dt_orig = datetime.datetime.fromisoformat(iso_string.replace('Z',''))
            return dt_orig.replace(tzinfo=None)
        except:
            return iso_string

def generate_excel_report(sklop_info, greske_data, report_title, is_global_report=False):
    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:30]
    font_calibri_bold_14 = Font(name='Calibri', size=14, bold=True); font_calibri_bold_10 = Font(name='Calibri', size=10, bold=True); font_calibri_regular_11 = Font(name='Calibri', size=11); font_calibri_regular_9 = Font(name='Calibri', size=9); font_calibri_green_11 = Font(name='Calibri', size=11, color="008000", bold=True); font_calibri_red_11 = Font(name='Calibri', size=11, color="FF0000", bold=True); center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); right_alignment = Alignment(horizontal='right', vertical='center'); header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"); thin_border_side = Side(border_style="thin", color="AAAAAA"); thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    logo_path = os.path.abspath(os.path.join(app.root_path, 'static', 'images', 'logo.png'))
    if os.path.exists(logo_path):
        try: img = OpenpyxlImage(logo_path); img.height = 30; original_height_guess = 42.5; img.width = (img.width * 30) / original_height_guess if original_height_guess > 0 else img.width; ws.add_image(img, 'A1'); ws.row_dimensions[1].height = 25
        except Exception as e: print(f"Greška dodavanja loga Excel: {e}")
    last_col_letter_header = 'K' if is_global_report else 'H'; ws.merge_cells(f'D1:{last_col_letter_header}1'); header_cell = ws['D1']; header_cell.value = f"Projekt: {sklop_info.get('projekt_naziv', 'N/A')} | Sklop: {sklop_info.get('sklop_naziv', 'N/A')}"; header_cell.font = font_calibri_regular_9; header_cell.alignment = right_alignment
    ws.merge_cells(f'A3:{last_col_letter_header}3'); title_cell = ws['A3']; title_cell.value = report_title; title_cell.font = font_calibri_bold_14; title_cell.alignment = center_alignment; ws.row_dimensions[3].height = 25
    ws.merge_cells(f'A4:{last_col_letter_header}4'); subtitle_cell = ws['A4']; subtitle_cell.value = f"Projekt: {sklop_info.get('projekt_naziv','N/A')} | Sklop: {sklop_info.get('sklop_naziv','N/A')} (ID: {sklop_info.get('sklop_id','N/A') if sklop_info.get('sklop_id') else 'Svi'})"; subtitle_cell.font = font_calibri_regular_11; subtitle_cell.alignment = center_alignment
    start_row_table = 6
    if is_global_report: headers = ['Projekt', 'Kupac', 'Sklop', 'Rb.', 'Tip Greške', 'Opis Greške', 'Evidentirao', 'Vrijeme Evid.', 'Status', 'Riješio', 'Vrijeme Rješ.']; col_widths_excel = {'A': 25, 'B': 15, 'C': 25, 'D': 6, 'E': 25, 'F': 40, 'G': 15, 'H': 18, 'I': 10, 'J': 15, 'K': 18}
    else: headers = ["Rb.", "Tip Greške", "Opis Greške", "Evidentirao", "Vrijeme Evid.", "Status", "Riješio", "Vrijeme Rješ."]; col_widths_excel = {'A': 5, 'B': 30, 'C': 55, 'D': 20, 'E': 20, 'F': 15, 'G': 20, 'H': 20}
    for i, header in enumerate(headers):
        col_letter = get_column_letter(i + 1); cell = ws[f"{col_letter}{start_row_table}"]; cell.value = header; cell.font = font_calibri_bold_10; cell.alignment = center_alignment; cell.fill = header_fill; cell.border = thin_border
        if col_letter in col_widths_excel: ws.column_dimensions[col_letter].width = col_widths_excel[col_letter]
    def format_first_word_title_case(text):
        if not text or not isinstance(text, str): return text
        words = text.split(' ', 1); first_word_capitalized = words[0].capitalize(); return first_word_capitalized + (' ' + words[1] if len(words) > 1 else '')
    current_row = start_row_table + 1; rb_counter = 1
    if not greske_data: last_col_letter_data = get_column_letter(len(headers)); ws.merge_cells(f'A{current_row}:{last_col_letter_data}{current_row}'); cell = ws[f'A{current_row}']; cell.value = 'Nema podataka.'; cell.font = font_calibri_regular_11; cell.alignment = center_alignment; cell.border = thin_border
    else:
        for row_data in greske_data:
            row = dict(row_data); vrijeme_evid_dt = format_datetime_hr_excel(row.get("VrijemeEvidentiranja")); vrijeme_rjes_dt = None; status_text = "Nije riješeno"; status_font = font_calibri_red_11
            if row.get('Rijesio'): vrijeme_rjes_dt = format_datetime_hr_excel(row.get("VrijemeRjesenja")); status_text = "Riješeno"; status_font = font_calibri_green_11
            col_idx = 1
            if is_global_report: ws[f'{get_column_letter(col_idx)}{current_row}'] = row.get('Projekt', 'N/A'); col_idx += 1; ws[f'{get_column_letter(col_idx)}{current_row}'] = row.get('Kupac', 'N/A'); col_idx += 1; ws[f'{get_column_letter(col_idx)}{current_row}'] = row.get('Sklop', 'N/A'); col_idx += 1
            ws[f'{get_column_letter(col_idx)}{current_row}'] = rb_counter; col_idx += 1
            tip_greske_formatted = format_first_word_title_case(row.get('TipGreske', 'N/A')); opis_greske_formatted = format_first_word_title_case(row.get('OpisGreske', 'N/A'))
            ws[f'{get_column_letter(col_idx)}{current_row}'] = tip_greske_formatted; col_idx += 1; ws[f'{get_column_letter(col_idx)}{current_row}'] = opis_greske_formatted; col_idx += 1
            ws[f'{get_column_letter(col_idx)}{current_row}'] = row.get('Evidentirao', 'N/A'); col_idx += 1
            cell_ve = ws[f'{get_column_letter(col_idx)}{current_row}']; cell_ve.value = vrijeme_evid_dt if isinstance(vrijeme_evid_dt, datetime.datetime) else None; cell_ve.number_format = 'dd.mm.yyyy. hh:mm:ss' if isinstance(vrijeme_evid_dt, datetime.datetime) else '@'; col_idx += 1
            cell_st = ws[f'{get_column_letter(col_idx)}{current_row}']; cell_st.value = status_text; cell_st.font = status_font; col_idx += 1
            ws[f'{get_column_letter(col_idx)}{current_row}'] = row.get('Rijesio', '-'); col_idx += 1
            cell_vr = ws[f'{get_column_letter(col_idx)}{current_row}']; cell_vr.value = vrijeme_rjes_dt if isinstance(vrijeme_rjes_dt, datetime.datetime) else "-"; cell_vr.number_format = 'dd.mm.yyyy. hh:mm:ss' if isinstance(vrijeme_rjes_dt, datetime.datetime) else '@'; col_idx += 1
            for i in range(1, col_idx):
                col_letter = get_column_letter(i); cell = ws[f"{col_letter}{current_row}"]; cell.border = thin_border; cell.alignment = left_alignment
                if not ((is_global_report and col_letter == 'I') or (not is_global_report and col_letter == 'F')): cell.font = font_calibri_regular_11
                if (is_global_report and col_letter == 'D') or (not is_global_report and col_letter == 'A'): cell.alignment = center_alignment
                if (is_global_report and col_letter == 'I') or (not is_global_report and col_letter == 'F'): cell.alignment = center_alignment
                if col_letter in ['H', 'K'] if is_global_report else col_letter in ['E', 'H']: cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            rb_counter += 1; current_row += 1
    try:
        ws.header_footer.odd_footer.left.text = "Zapisnik kreirao: Kontrolor _________"; ws.header_footer.odd_footer.left.size = 9; ws.header_footer.odd_footer.left.font = "Calibri,Regular"
        ws.header_footer.odd_footer.right.text = "Generirano: &[Date] &[Time]"; ws.header_footer.odd_footer.right.size = 9; ws.header_footer.odd_footer.right.font = "Calibri,Regular"
        ws.print_options.horizontalCentered = True
    except AttributeError: print("UPOZORENJE: Nije moguće postaviti Print Footer.")
    output_buffer = BytesIO()
    try:
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer
    except Exception as e: print(f"Greška kod spremanja Excela: {e}"); return None

# ISPRAVLJENA FUNKCIJA
def generate_pdf_report(html_content):
    pdf_buffer = BytesIO()
    try:
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer, encoding='utf-8', link_callback=fetch_resources)
        if pisa_status.err:
            print(f"!!! xhtml2pdf GREŠKA U PISA STATUSU: {pisa_status.err}")
            return None
        pdf_buffer.seek(0)
        return pdf_buffer
    except Exception as e:
        print(f"!!! DETALJNA GREŠKA KOD GENERIRANJA PDF-a: {e}")
        return None

# VRAĆENA ISPRAVNA, ROBUSTNA VERZIJA FUNKCIJE
def fetch_resources(uri, rel):
    static_folder_path = os.path.join(app.root_path, 'static')
    path = ""
    if uri.startswith('/static/'):
        path = os.path.join(static_folder_path, uri.split('/static/', 1)[1])
    elif uri.startswith('static/'):
        path = os.path.join(static_folder_path, uri.split('static/', 1)[1])
    else:
        path = os.path.join(static_folder_path, uri)
    
    if not os.path.abspath(path).startswith(os.path.abspath(static_folder_path)):
        return None
    if os.path.exists(path):
        return path
    return None

def create_html_for_pdf(sklop_info, greske_data, report_title, ime_korisnika="Nepoznat korisnik"):
    logo_tag = ""; logo_uri = None
    try:
        logo_file_path = os.path.join(app.static_folder, 'images', 'logo.png')
        if os.path.exists(logo_file_path): logo_uri = '/static/images/logo.png'
    except Exception as e: print(f"Greška kod provjere loga za PDF: {e}")
    if logo_uri: logo_tag = f'<img src="{logo_uri}" style="height: 30px; width: auto; vertical-align: middle;" alt="Logo"/>'
    
    font_face_rule = ""
    try:
        font_path = os.path.join(app.static_folder, 'fonts', 'aptos.ttf')
        if os.path.exists(font_path):
            with open(font_path, "rb") as font_file:
                font_data = font_file.read()
            font_base64 = base64.b64encode(font_data).decode('utf-8')
            font_face_rule = f"""@font-face {{ font-family: "aptos"; src: url(data:font/truetype;charset=utf-8;base64,{font_base64}); }}"""
        else:
            print("!!! UPOZORENJE: Font aptos.ttf nije pronađen u static/fonts folderu.")
    except Exception as e:
        print(f"!!! GREŠKA kod učitavanja fonta za PDF: {e}")

    styles = f"""<style>
        {font_face_rule}
        @page {{ size: A4 landscape; margin: 1cm; @frame header_frame {{ -pdf-frame-content: header_content; left: 1cm; width: 27.7cm; top: 0.5cm; height: 2.5cm; }} @frame content_frame {{ left: 1cm; width: 27.7cm; top: 3.2cm; height: 16.3cm; }} @frame footer_frame {{ -pdf-frame-content: footer_content; left: 1cm; width: 27.7cm; bottom: 0.5cm; height: 1cm; }} }}
        * {{ font-family: "aptos", sans-serif; }} body {{ font-size: 8pt; }} #header_content {{ font-size: 8pt; color: #333; }} #header_content .logo {{ float: left; margin-right: 10px; }} #header_content .project-info {{ text-align: right; }} h1 {{ text-align: center; color: #333; font-size: 14pt; margin: 5px 0; }} table {{ width: 100%; border-collapse: collapse; margin-top: 5px; table-layout: fixed; }} th, td {{ border: 1px solid #bbb; padding: 3px 4px; text-align: left; vertical-align: top; word-wrap: break-word; }} th {{ background-color: #e8e8e8; font-weight: bold; text-align: center; }} .nije-rijeseno {{ color: red; font-weight: bold; }} .rijeseno {{ color: green; }} #footer_content {{ text-align: right; font-size: 8pt; color: #777; }} .row-even {{ background-color: #f9f9f9; }} td.col-rb {{ text-align: center; width: 4%; }} td.col-tip {{ width: 18%; }} td.col-opis {{ width: 28%; }} td.col-evidentirao {{ width: 12%; }} td.col-vrijeme-evid {{ width: 12%; }} td.col-status {{ width: 8%; text-align: center;}} td.col-rijesio {{ width: 8%; }} td.col-vrijeme-rjes {{ width: 10%; }}
    </style>"""
    def format_first_word_title_case(text):
        if not text or not isinstance(text, str): return text
        words = text.split(' ', 1); return words[0].capitalize() + (' ' + words[1] if len(words) > 1 else '')
    table_rows = ""; rb = 1
    for i, row_data in enumerate(greske_data):
        row = dict(row_data); dt_evid = format_datetime_hr_excel(row.get("VrijemeEvidentiranja"), target_timezone='Europe/Zagreb'); vrijeme_evid_str = dt_evid.strftime('%d.%m.%y %H:%M') if isinstance(dt_evid, datetime.datetime) else "N/A"
        dt_rjes = format_datetime_hr_excel(row.get("VrijemeRjesenja"), target_timezone='Europe/Zagreb'); vrijeme_rjes_str = dt_rjes.strftime('%d.%m.%y %H:%M') if isinstance(dt_rjes, datetime.datetime) else "-"
        status_html = '<span class="nije-rijeseno">Nije riješeno</span>'
        if row.get('Rijesio'): status_html = f'<span class="rijeseno">Riješeno</span>'
        tip_greske_formatted = format_first_word_title_case(row.get('TipGreske', 'N/A')); opis_greske_formatted = format_first_word_title_case(row.get('OpisGreske', 'N/A'))
        row_class = "row-even" if i % 2 == 0 else "row-odd"
        table_rows += f""" <tr class="{row_class}"> <td class="col-rb">{rb}</td> <td class="col-tip">{tip_greske_formatted}</td> <td class="col-opis">{opis_greske_formatted}</td> <td class="col-evidentirao">{row.get('Evidentirao', 'N/A')}</td> <td class="col-vrijeme-evid">{vrijeme_evid_str}</td> <td class="col-status">{status_html}</td> <td class="col-rijesio">{row.get('Rijesio', '-')}</td> <td class="col-vrijeme-rjes">{vrijeme_rjes_str}</td> </tr> """
        rb += 1
    if not greske_data: table_rows = '<tr><td colspan="8" style="text-align:center; padding: 15px;">Nema podataka.</td></tr>'
    html = f""" <!DOCTYPE html> <html lang="hr"> <head> <meta charset="UTF-8"> <title>{report_title}</title> {styles} </head> <body> <div id="header_content"> <div class="logo">{logo_tag}</div> <div class="project-info"> Projekt: <strong>{sklop_info.get('projekt_naziv','N/A')}</strong><br/> Sklop: <strong>{sklop_info.get('sklop_naziv','N/A')}</strong> (ID: {sklop_info.get('sklop_id','N/A')}) </div> <div style="clear: both;"></div> </div> <h1>{report_title}</h1> <table> <thead> <tr> <th>Rb.</th> <th>Tip Greške</th> <th>Opis Greške</th> <th>Evidentirao</th> <th>Vrijeme Evid.</th> <th>Status</th> <th>Riješio</th> <th>Vrijeme Rješ.</th> </tr> </thead> <tbody> {table_rows} </tbody> </table> <div id="footer_content"> Zapisnik kreirao: {ime_korisnika} | Generirano: {datetime.datetime.now().strftime('%d.%m.%Y. %H:%M:%S')} | Stranica <pdf:pagenumber> od <pdf:pagecount> </div> </body> </html> """
    return html

# === 4. Glavne Rute Aplikacije ===
@app.route('/')
def home_redirect():
    if 'username' in session: return redirect(url_for('index'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'username' in session: return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = USERS.get(username)
        if user and user['password'] == password:
            session['username'] = username
            flash('Uspješno ste prijavljeni!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Pogrešno korisničko ime ili lozinka.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Uspješno ste odjavljeni.', 'info')
    return redirect(url_for('login'))

@app.route('/index')
@login_required
def index():
    db = get_db()
    cursor = db.execute('SELECT id, naziv FROM projekti ORDER BY naziv COLLATE NOCASE')
    projekti = cursor.fetchall()
    logo_url = None
    try:
        logo_file_path = os.path.join(app.static_folder, 'images', 'logo.png')
        if os.path.exists(logo_file_path):
            logo_url = url_for('static', filename='images/logo.png')
    except Exception as e:
        print(f"Greška kod dohvaćanja URL-a za logo: {e}")
    return render_template('index.html', projekti=projekti, logo_url=logo_url)

@app.route('/uploads/greske/<path:filename>')
@login_required
def uploaded_file_greska(filename):
    safe_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename))
    if not safe_path.startswith(os.path.abspath(app.config['UPLOAD_FOLDER_GRESKE'])):
        return "Zabranjeno", 403
    return send_from_directory(app.config['UPLOAD_FOLDER_GRESKE'], filename)

# === 5. API Rute (Sve zaštićene) ===
@app.route('/api/projekti', methods=['GET'])
@login_required
def api_dohvati_projekte():
    kupac_filter = request.args.get('kupac')
    db = get_db(); sql = 'SELECT id, naziv, kupac FROM projekti'; params = []
    if kupac_filter and kupac_filter != 'Svi': sql += ' WHERE kupac = ?'; params.append(kupac_filter)
    sql += ' ORDER BY naziv COLLATE NOCASE'; cursor = db.execute(sql, params); projekti = cursor.fetchall()
    return jsonify([dict(p) for p in projekti])

@app.route('/api/kupci', methods=['GET'])
@login_required
def api_dohvati_kupce():
    db = get_db(); cursor = db.execute('SELECT DISTINCT kupac FROM projekti WHERE kupac IS NOT NULL AND kupac != "" ORDER BY kupac COLLATE NOCASE')
    kupci = [row['kupac'] for row in cursor.fetchall()]; return jsonify(kupci)

@app.route('/api/projekti', methods=['POST'])
@login_required
def api_dodaj_projekt():
    data = request.get_json(); ime_projekta = data.get('ime', '').strip(); kupac = data.get('kupac', '').strip()
    if not ime_projekta or not kupac: return jsonify({"greska": "Ime projekta i Kupac su obavezni"}), 400
    db = get_db()
    try: cursor = db.execute('INSERT INTO projekti (naziv, kupac) VALUES (?, ?)', (ime_projekta, kupac)); db.commit(); novi_id = cursor.lastrowid; return jsonify({"id": novi_id, "ime": ime_projekta, "kupac": kupac}), 201
    except sqlite3.IntegrityError: db.rollback(); return jsonify({"greska": f"Projekt '{ime_projekta}' već postoji."}), 409
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska":"Greška u bazi podataka kod dodavanja projekta."}), 500

@app.route('/api/projekti/<int:projekt_id>', methods=['DELETE'])
@login_required
def api_obrisi_projekat(projekt_id):
    data = request.get_json(); provided_password = data.get('password') if data else None
    if not provided_password: return jsonify({"greska": "Lozinka za brisanje je obavezna."}), 401
    if provided_password != DELETE_PASSWORD: return jsonify({"greska": "Pogrešna lozinka."}), 403
    db = get_db()
    try:
        sql_slike = "SELECT gs.naziv_datoteke FROM greska_slike gs JOIN greske g ON gs.greska_id = g.id JOIN sklopovi s ON g.sklop_id = s.id WHERE s.projekt_id = ?"
        cursor_slike = db.execute(sql_slike, (projekt_id,)); slike_za_brisanje = [row['naziv_datoteke'] for row in cursor_slike.fetchall()]
        cursor_delete = db.execute('DELETE FROM projekti WHERE id = ?', (projekt_id,))
        if cursor_delete.rowcount == 0: return jsonify({"greska": f"Projekat ID {projekt_id} nije pronađen."}), 404
        db.commit()
        for filename in slike_za_brisanje:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename)
                if os.path.exists(filepath): os.remove(filepath)
            except Exception as e: print(f"Greška kod brisanja datoteke slike {filename}: {e}")
        return jsonify({"message": f"Projekat ID {projekt_id} i svi povezani podaci su uspješno obrisani."}), 200
    except Exception as e: db.rollback(); return jsonify({"greska": f"Greška kod brisanja projekta: {e}"}), 500

@app.route('/api/sklopovi/<int:projekt_id>')
@login_required
def api_dohvati_sklopove(projekt_id):
    db = get_db(); cursor = db.execute('SELECT id, naziv FROM sklopovi WHERE projekt_id = ? ORDER BY naziv COLLATE NOCASE', (projekt_id,)); sklopovi = cursor.fetchall()
    return jsonify([dict(s) for s in sklopovi])

@app.route('/api/sklopovi', methods=['POST'])
@login_required
def api_dodaj_sklop():
    data = request.get_json(); ime_sklopa = data.get('ime', '').strip(); projekt_id = data.get('projekt_id')
    if not ime_sklopa or not projekt_id: return jsonify({"greska": "Ime sklopa i ID projekta su obavezni."}), 400
    db = get_db(); cursor_projekt = db.execute('SELECT id FROM projekti WHERE id = ?', (projekt_id,));
    if cursor_projekt.fetchone() is None: return jsonify({"greska": f"Projekt ID {projekt_id} ne postoji."}), 404
    cursor_check = db.execute('SELECT id FROM sklopovi WHERE projekt_id = ? AND lower(naziv) = lower(?)', (projekt_id, ime_sklopa))
    if cursor_check.fetchone() is not None: return jsonify({"greska": f"Sklop '{ime_sklopa}' već postoji u ovom projektu."}), 409
    try: cursor = db.execute('INSERT INTO sklopovi (projekt_id, naziv) VALUES (?, ?)', (projekt_id, ime_sklopa)); db.commit(); novi_id = cursor.lastrowid; return jsonify({"id": novi_id, "naziv": ime_sklopa, "projekt_id": projekt_id}), 201
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska":"Greška u bazi podataka kod dodavanja sklopa."}), 500

@app.route('/api/sklopovi/<int:sklop_id>', methods=['PUT'])
@login_required
def api_preimenuj_sklop(sklop_id):
    data = request.get_json(); novo_ime = data.get('naziv', '').strip()
    if not novo_ime: return jsonify({"greska": "Novo ime sklopa je obavezno."}), 400
    db = get_db(); cursor_check = db.execute('SELECT projekt_id FROM sklopovi WHERE id = ?', (sklop_id,)); sklop = cursor_check.fetchone()
    if sklop is None: return jsonify({"greska": f"Sklop ID {sklop_id} nije pronađen."}), 404
    projekt_id = sklop['projekt_id']; cursor_duplikat = db.execute('SELECT id FROM sklopovi WHERE projekt_id = ? AND lower(naziv) = lower(?) AND id != ?', (projekt_id, novo_ime, sklop_id))
    if cursor_duplikat.fetchone() is not None: return jsonify({"greska": f"Sklop '{novo_ime}' već postoji u ovom projektu."}), 409
    try: db.execute('UPDATE sklopovi SET naziv = ? WHERE id = ?', (novo_ime, sklop_id)); db.commit(); return jsonify({"id": sklop_id, "naziv": novo_ime, "projekt_id": projekt_id}), 200
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska": "Greška u bazi podataka kod preimenovanja sklopa."}), 500

@app.route('/api/sklopovi/<int:sklop_id>', methods=['DELETE'])
@login_required
def api_obrisi_sklop(sklop_id):
    data = request.get_json(); provided_password = data.get('password') if data else None
    if not provided_password: return jsonify({"greska": "Lozinka za brisanje je obavezna."}), 401
    if provided_password != DELETE_PASSWORD: return jsonify({"greska": "Pogrešna lozinka."}), 403
    db = get_db()
    try:
        cursor_slike = db.execute('SELECT gs.naziv_datoteke FROM greska_slike gs JOIN greske g ON gs.greska_id = g.id WHERE g.sklop_id = ?', (sklop_id,)); slike_za_brisanje = [row['naziv_datoteke'] for row in cursor_slike.fetchall()]
        cursor_delete = db.execute('DELETE FROM sklopovi WHERE id = ?', (sklop_id,))
        if cursor_delete.rowcount == 0: return jsonify({"greska": f"Sklop ID {sklop_id} nije pronađen."}), 404
        db.commit()
        for filename in slike_za_brisanje:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename)
                if os.path.exists(filepath): os.remove(filepath)
            except Exception as e: print(f"Greška kod brisanja datoteke slike {filename}: {e}")
        return jsonify({"message": f"Sklop {sklop_id} i sve povezane greške uspješno obrisani."}), 200
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska": "Greška u bazi podataka kod brisanja sklopa."}), 500

@app.route('/api/greske/sklop/<int:sklop_id>')
@login_required
def api_dohvati_greske_za_sklop(sklop_id):
    db = get_db(); cursor_sklop = db.execute('SELECT id FROM sklopovi WHERE id = ?', (sklop_id,))
    if cursor_sklop.fetchone() is None: return jsonify([]), 200
    cursor = db.execute('SELECT g.*, GROUP_CONCAT(gs.naziv_datoteke) AS slike_datoteke FROM greske g LEFT JOIN greska_slike gs ON g.id = gs.greska_id WHERE g.sklop_id = ? GROUP BY g.id ORDER BY g.id ASC', (sklop_id,)); greske = cursor.fetchall(); greske_lista = []
    for g in greske: greska_dict = dict(g); slike_str = greska_dict.get('slike_datoteke'); greska_dict['slike'] = slike_str.split(',') if slike_str else []; del greska_dict['slike_datoteke']; greske_lista.append(greska_dict)
    return jsonify(greske_lista)

@app.route('/api/greske', methods=['POST'])
@login_required
def api_dodaj_gresku():
    sklop_id = request.form.get('sklop_id', type=int); tip_greske = request.form.get('tip', '').strip(); opis_greske = request.form.get('mjesto', '').strip(); evidentirao = request.form.get('evidentirao', '').strip()
    if not all([sklop_id, tip_greske, opis_greske, evidentirao]): return jsonify({"greska": "Tip, Opis i Evidentirao su obavezni."}), 400
    db = get_db(); cursor_sklop = db.execute('SELECT id FROM sklopovi WHERE id = ?', (sklop_id,))
    if cursor_sklop.fetchone() is None: return jsonify({"greska": f"Sklop ID {sklop_id} ne postoji."}), 404
    sada_iso = datetime.datetime.now(datetime.timezone.utc).isoformat(); greska_id = None; spremljene_slike_nazivi = []
    try:
        cursor = db.execute('INSERT INTO greske (sklop_id, tip, mjesto, vrijeme_evidentiranja, evidentirao) VALUES (?, ?, ?, ?, ?)', (sklop_id, tip_greske, opis_greske, sada_iso, evidentirao)); greska_id = cursor.lastrowid
        uploaded_files = request.files.getlist('slike[]'); sada_upload_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
        for file in uploaded_files:
            if file and allowed_file(file.filename):
                original_filename = secure_filename(file.filename); file_ext = os.path.splitext(original_filename)[1]; unique_filename = f"{uuid.uuid4().hex}{file_ext}"; save_path = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], unique_filename)
                try: file.save(save_path); db.execute('INSERT INTO greska_slike (greska_id, naziv_datoteke, originalni_naziv, vrijeme_uploada) VALUES (?, ?, ?, ?)', (greska_id, unique_filename, original_filename, sada_upload_iso)); spremljene_slike_nazivi.append(unique_filename)
                except Exception as upload_err: print(f"Greška kod obrade slike '{original_filename}': {upload_err}")
        db.commit()
        cursor_new = db.execute('SELECT g.*, GROUP_CONCAT(gs.naziv_datoteke) AS slike_datoteke FROM greske g LEFT JOIN greska_slike gs ON g.id = gs.greska_id WHERE g.id = ? GROUP BY g.id', (greska_id,)); nova_greska_row = cursor_new.fetchone()
        if nova_greska_row: greska_dict = dict(nova_greska_row); slike_str = greska_dict.get('slike_datoteke'); greska_dict['slike'] = slike_str.split(',') if slike_str else []; del greska_dict['slike_datoteke']; return jsonify(greska_dict), 201
        else: return jsonify({"greska":"Greška spremljena, ali došlo je do problema kod dohvaćanja potvrde."}), 500
    except Exception as e:
        db.rollback()
        for filename in spremljene_slike_nazivi:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename)
                if os.path.exists(filepath): os.remove(filepath)
            except OSError as remove_err: print(f"Nije moguće obrisati sliku {filename} tijekom rollbacka: {remove_err}")
        return jsonify({"greska": f"Došlo je do greške: {e}"}), 500

@app.route('/api/greske/<int:greska_id>', methods=['GET'])
@login_required
def api_dohvati_gresku(greska_id):
    db = get_db(); cursor = db.execute('SELECT * FROM greske WHERE id = ?', (greska_id,)); greska = cursor.fetchone()
    if greska is None: return jsonify({"greska": f"Greška ID {greska_id} nije pronađena."}), 404
    cursor_slike = db.execute('SELECT naziv_datoteke FROM greska_slike WHERE greska_id = ? ORDER BY id', (greska_id,)); slike = [row['naziv_datoteke'] for row in cursor_slike.fetchall()]
    greska_dict = dict(greska); greska_dict['slike'] = slike; return jsonify(greska_dict)

@app.route('/api/greske/<int:greska_id>', methods=['PUT'])
@login_required
def api_uredi_gresku(greska_id):
    db = get_db(); cursor_check = db.execute('SELECT * FROM greske WHERE id = ?', (greska_id,)); postojeca_greska = cursor_check.fetchone()
    if postojeca_greska is None: return jsonify({"greska": f"Greška ID {greska_id} nije pronađena."}), 404
    tip = request.form.get('tip', postojeca_greska['tip']).strip(); opis_greske = request.form.get('mjesto', postojeca_greska['mjesto']).strip(); rijesio_form = request.form.get('rijesio')
    if not tip or not opis_greske: return jsonify({"greska": "Tip i Opis greške su obavezni."}), 400
    rijesio_db_value = postojeca_greska['rijesio']; vrijeme_rjesenja_db_value = postojeca_greska['vrijeme_rjesenja']; update_solved_status = False; new_rijesio_value = rijesio_db_value; new_vrijeme_rjesenja_value = vrijeme_rjesenja_db_value
    if rijesio_form is not None:
        rijesio_form_strip = rijesio_form.strip()
        if rijesio_form_strip and rijesio_db_value != rijesio_form_strip: new_rijesio_value = rijesio_form_strip; new_vrijeme_rjesenja_value = datetime.datetime.now(datetime.timezone.utc).isoformat(); update_solved_status = True
        elif not rijesio_form_strip and rijesio_db_value is not None: new_rijesio_value = None; new_vrijeme_rjesenja_value = None; update_solved_status = True
    spremljene_slike_nove = [];
    try:
        sql = 'UPDATE greske SET tip = ?, mjesto = ?'; params = [tip, opis_greske]
        if update_solved_status: sql += ', rijesio = ?, vrijeme_rjesenja = ?'; params.extend([new_rijesio_value, new_vrijeme_rjesenja_value])
        sql += ' WHERE id = ?'; params.append(greska_id); db.execute(sql, tuple(params))
        uploaded_files = request.files.getlist('slike_edit[]'); sada_upload_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
        for file in uploaded_files:
            if file and allowed_file(file.filename):
                original_filename = secure_filename(file.filename); file_ext = os.path.splitext(original_filename)[1]; unique_filename = f"{uuid.uuid4().hex}{file_ext}"; save_path = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], unique_filename)
                try: file.save(save_path); db.execute('INSERT INTO greska_slike (greska_id, naziv_datoteke, originalni_naziv, vrijeme_uploada) VALUES (?, ?, ?, ?)', (greska_id, unique_filename, original_filename, sada_upload_iso)); spremljene_slike_nove.append(unique_filename)
                except Exception as upload_err: print(f"Greška kod dodavanja NOVE slike: {upload_err}")
        slike_za_brisanje_nazivi = request.form.getlist('obrisi_sliku[]')
        for filename_to_delete in slike_za_brisanje_nazivi:
            safe_filename = secure_filename(filename_to_delete)
            if safe_filename:
                file_path_to_delete = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], safe_filename)
                try:
                    cursor_delete = db.execute('DELETE FROM greska_slike WHERE greska_id = ? AND naziv_datoteke = ?', (greska_id, safe_filename))
                    if cursor_delete.rowcount > 0 and os.path.exists(file_path_to_delete):
                        os.remove(file_path_to_delete)
                except Exception as delete_err: print(f"Greška kod brisanja slike: {delete_err}")
        db.commit()
        cursor_updated = db.execute('SELECT g.*, GROUP_CONCAT(gs.naziv_datoteke) AS slike_datoteke FROM greske g LEFT JOIN greska_slike gs ON g.id = gs.greska_id WHERE g.id = ? GROUP BY g.id', (greska_id,)); azurirana_greska = cursor_updated.fetchone()
        if azurirana_greska: greska_dict = dict(azurirana_greska); slike_str = greska_dict.get('slike_datoteke'); greska_dict['slike'] = slike_str.split(',') if slike_str else []; del greska_dict['slike_datoteke']; return jsonify(greska_dict), 200
        else: return jsonify({"greska":"Greška ažurirana, ali problem kod dohvaćanja potvrde."}), 500
    except Exception as e:
        db.rollback()
        for filename in spremljene_slike_nove:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename)
                if os.path.exists(filepath): os.remove(filepath)
            except OSError as remove_err: print(f"Nije moguće obrisati novu sliku {filename}: {remove_err}")
        return jsonify({"greska": f"Došlo je do neočekivane greške: {e}"}), 500

@app.route('/api/greske/<int:greska_id>/rijesi', methods=['POST'])
@login_required
def api_rijesi_gresku(greska_id):
    data = request.get_json(); rijesio_ime = data.get('rijesio', '').strip()
    if not rijesio_ime: return jsonify({"greska": "Ime rješavatelja je obavezno."}), 400
    db = get_db(); cursor_check = db.execute('SELECT rijesio FROM greske WHERE id = ?', (greska_id,)); greska_status = cursor_check.fetchone()
    if greska_status is None: return jsonify({"greska": f"Greška ID {greska_id} nije pronađena."}), 404
    if greska_status['rijesio'] is not None and greska_status['rijesio'] != '': return jsonify({"greska": f"Greška {greska_id} je već riješena."}), 400
    sada = datetime.datetime.now(datetime.timezone.utc).isoformat()
    try:
        db.execute('UPDATE greske SET rijesio = ?, vrijeme_rjesenja = ? WHERE id = ?', (rijesio_ime, sada, greska_id)); db.commit()
        cursor_updated = db.execute('SELECT * FROM greske WHERE id = ?', (greska_id,)); azurirana_greska = cursor_updated.fetchone()
        if azurirana_greska:
             cursor_slike = db.execute('SELECT naziv_datoteke FROM greska_slike WHERE greska_id = ? ORDER BY id', (greska_id,)); slike = [row['naziv_datoteke'] for row in cursor_slike.fetchall()]
             greska_dict = dict(azurirana_greska); greska_dict['slike'] = slike; return jsonify(greska_dict), 200
        else: return jsonify({"greska":"Status ažuriran, ali greška kod dohvaćanja potvrde."}), 500
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska":"Greška u bazi podataka kod označavanja greške kao riješene."}), 500

@app.route('/api/greske/<int:greska_id>', methods=['DELETE'])
@login_required
def api_obrisi_gresku(greska_id):
    data = request.get_json(); provided_password = data.get('password') if data else None
    if not provided_password: return jsonify({"greska": "Lozinka za brisanje je obavezna."}), 401
    if provided_password != DELETE_PASSWORD: return jsonify({"greska": "Pogrešna lozinka."}), 403
    db = get_db()
    try:
        cursor_slike = db.execute('SELECT naziv_datoteke FROM greska_slike WHERE greska_id = ?', (greska_id,)); slike_za_brisanje = [row['naziv_datoteke'] for row in cursor_slike.fetchall()]
        cursor_delete = db.execute('DELETE FROM greske WHERE id = ?', (greska_id,))
        if cursor_delete.rowcount == 0: return jsonify({"greska": f"Greška ID {greska_id} nije pronađena."}), 404
        db.commit()
        for filename in slike_za_brisanje:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER_GRESKE'], filename)
                if os.path.exists(filepath): os.remove(filepath)
            except Exception as e: print(f"Greška kod brisanja slike {filename}: {e}")
        return jsonify({"message": f"Greška {greska_id} uspješno obrisana."}), 200
    except sqlite3.Error as e: db.rollback(); return jsonify({"greska": "Greška u bazi podataka kod brisanja greške."}), 500

# === 6. Export Rute (ZAŠTIĆENE) ===
@app.route('/api/greske/sklop/<int:sklop_id>/excel/sve')
@login_required
def export_excel_sklop_sve(sklop_id):
    db = get_db(); cursor_sklop = db.execute('SELECT s.id as sklop_id, s.naziv as sklop_naziv, p.naziv as projekt_naziv, p.kupac as kupac FROM sklopovi s JOIN projekti p ON s.projekt_id = p.id WHERE s.id = ?', (sklop_id,)); sklop_info = cursor_sklop.fetchone()
    if not sklop_info: return "Sklop nije pronađen", 404
    cursor_greske = db.execute('SELECT id as GreskaID, tip as TipGreske, mjesto as OpisGreske, evidentirao as Evidentirao, vrijeme_evidentiranja as VrijemeEvidentiranja, rijesio as Rijesio, vrijeme_rjesenja as VrijemeRjesenja FROM greske WHERE sklop_id = ? ORDER BY id ASC', (sklop_id,)); greske = cursor_greske.fetchall()
    excel_buffer = generate_excel_report(dict(sklop_info), greske, f"Izvještaj svih grešaka - Sklop {sklop_info['sklop_naziv']}", is_global_report=False)
    if excel_buffer is None: return "Greška pri generiranju Excela.", 500
    filename = f"{sklop_info['projekt_naziv']}_izvjestaj_svih_gresaka_sklop_{sklop_info['sklop_naziv']}_{datetime.date.today().strftime('%d.%m.%Y')}.xlsx"
    return Response(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment;filename="{filename}"'})

@app.route('/api/greske/sklop/<int:sklop_id>/excel/nerijesene')
@login_required
def export_excel_sklop_nerijesene(sklop_id):
    db = get_db(); cursor_sklop = db.execute('SELECT s.id as sklop_id, s.naziv as sklop_naziv, p.naziv as projekt_naziv, p.kupac as kupac FROM sklopovi s JOIN projekti p ON s.projekt_id = p.id WHERE s.id = ?', (sklop_id,)); sklop_info = cursor_sklop.fetchone()
    if not sklop_info: return "Sklop nije pronađen", 404
    cursor_greske = db.execute('SELECT id as GreskaID, tip as TipGreske, mjesto as OpisGreske, evidentirao as Evidentirao, vrijeme_evidentiranja as VrijemeEvidentiranja, rijesio as Rijesio, vrijeme_rjesenja as VrijemeRjesenja FROM greske WHERE sklop_id = ? AND (rijesio IS NULL OR rijesio = "") ORDER BY id ASC', (sklop_id,)); greske = cursor_greske.fetchall()
    excel_buffer = generate_excel_report(dict(sklop_info), greske, f"Izvještaj neriješenih grešaka - Sklop {sklop_info['sklop_naziv']}", is_global_report=False)
    if excel_buffer is None: return "Greška pri generiranju Excela.", 500
    filename = f"{sklop_info['projekt_naziv']}_izvjestaj_nerijesenih_gresaka_sklop_{sklop_info['sklop_naziv']}_{datetime.date.today().strftime('%d.%m.%Y')}.xlsx"
    return Response(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment;filename="{filename}"'})

@app.route('/api/greske/sklop/<int:sklop_id>/pdf/sve')
@login_required
def export_pdf_sklop_sve(sklop_id):
    db = get_db(); cursor_sklop = db.execute('SELECT s.id as sklop_id, s.naziv as sklop_naziv, p.naziv as projekt_naziv, p.kupac as kupac FROM sklopovi s JOIN projekti p ON s.projekt_id = p.id WHERE s.id = ?', (sklop_id,)); sklop_info = cursor_sklop.fetchone()
    if not sklop_info: return "Sklop nije pronađen", 404
    cursor_greske = db.execute('SELECT id as GreskaID, tip as TipGreske, mjesto as OpisGreske, evidentirao as Evidentirao, vrijeme_evidentiranja as VrijemeEvidentiranja, rijesio as Rijesio, vrijeme_rjesenja as VrijemeRjesenja FROM greske WHERE sklop_id = ? ORDER BY id ASC', (sklop_id,)); greske = cursor_greske.fetchall()
    ime_korisnika = USERS.get(session.get('username'), {}).get('full_name', 'Nepoznat')
    report_title = f"Izvještaj svih grešaka - Sklop {sklop_info['sklop_naziv']}"
    html_content = create_html_for_pdf(dict(sklop_info), greske, report_title, ime_korisnika)
    pdf_buffer = generate_pdf_report(html_content)
    if pdf_buffer is None: return "Greška pri generiranju PDF-a.", 500
    filename = f"{sklop_info['projekt_naziv']}_izvjestaj_svih_gresaka_sklop_{sklop_info['sklop_naziv']}_{datetime.date.today().strftime('%d.%m.%Y')}.pdf"
    return Response(pdf_buffer, mimetype='application/pdf', headers={'Content-Disposition': f'attachment;filename="{filename}"'})

@app.route('/api/greske/sklop/<int:sklop_id>/pdf/nerijesene')
@login_required
def export_pdf_sklop_nerijesene(sklop_id):
    db = get_db(); cursor_sklop = db.execute('SELECT s.id as sklop_id, s.naziv as sklop_naziv, p.naziv as projekt_naziv, p.kupac as kupac FROM sklopovi s JOIN projekti p ON s.projekt_id = p.id WHERE s.id = ?', (sklop_id,)); sklop_info = cursor_sklop.fetchone()
    if not sklop_info: return "Sklop nije pronađen", 404
    cursor_greske = db.execute('SELECT id as GreskaID, tip as TipGreske, mjesto as OpisGreske, evidentirao as Evidentirao, vrijeme_evidentiranja as VrijemeEvidentiranja, rijesio as Rijesio, vrijeme_rjesenja as VrijemeRjesenja FROM greske WHERE sklop_id = ? AND (rijesio IS NULL OR rijesio = "") ORDER BY id ASC', (sklop_id,)); greske = cursor_greske.fetchall()
    ime_korisnika = USERS.get(session.get('username'), {}).get('full_name', 'Nepoznat')
    report_title = f"Izvještaj neriješenih grešaka - Sklop {sklop_info['sklop_naziv']}"
    html_content = create_html_for_pdf(dict(sklop_info), greske, report_title, ime_korisnika)
    pdf_buffer = generate_pdf_report(html_content)
    if pdf_buffer is None: return "Greška pri generiranju PDF-a.", 500
    filename = f"{sklop_info['projekt_naziv']}_izvjestaj_nerijesenih_gresaka_sklop_{sklop_info['sklop_naziv']}_{datetime.date.today().strftime('%d.%m.%Y')}.pdf"
    return Response(pdf_buffer, mimetype='application/pdf', headers={'Content-Disposition': f'attachment;filename="{filename}"'})

@app.route('/export/excel')
@login_required
def export_excel_global():
    db = get_db(); query = """ SELECT p.naziv AS Projekt, p.kupac AS Kupac, s.naziv AS Sklop, g.id AS GreskaID, g.tip AS TipGreske, g.mjesto AS OpisGreske, g.evidentirao AS Evidentirao, g.vrijeme_evidentiranja AS VrijemeEvidentiranja, g.rijesio AS Rijesio, g.vrijeme_rjesenja AS VrijemeRjesenja FROM greske g JOIN sklopovi s ON g.sklop_id = s.id JOIN projekti p ON s.projekt_id = p.id ORDER BY p.naziv COLLATE NOCASE, s.naziv COLLATE NOCASE, g.id ASC; """; cursor = db.execute(query); data = cursor.fetchall()
    excel_buffer = generate_excel_report( {"projekt_naziv": "SVI", "sklop_naziv": "SVI", "sklop_id": None}, data, "Izvještaj svih grešaka (Globalni)", is_global_report=True )
    if excel_buffer is None: return "Greška pri generiranju Excela.", 500
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S"); filename = f"evidencija_gresaka_{timestamp}.xlsx"
    return Response( excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment;filename={filename}"} )

@app.route('/export/csv')
@login_required
def export_csv():
    db = get_db(); query = """ SELECT p.naziv AS Projekt, p.kupac AS Kupac, s.naziv AS Sklop, g.id AS GreskaID, g.tip AS TipGreske, g.mjesto AS OpisGreske, g.evidentirao AS Evidentirao, g.vrijeme_evidentiranja AS VrijemeEvidentiranja, g.rijesio AS Rijesio, g.vrijeme_rjesenja AS VrijemeRjesenja, GROUP_CONCAT(gs.originalni_naziv) AS SlikeOriginalniNazivi FROM greske g JOIN sklopovi s ON g.sklop_id = s.id JOIN projekti p ON s.projekt_id = p.id LEFT JOIN greska_slike gs ON g.id = gs.greska_id GROUP BY g.id ORDER BY p.naziv COLLATE NOCASE, s.naziv COLLATE NOCASE, g.id ASC; """; cursor = db.execute(query); data = cursor.fetchall()
    si = StringIO(); cw = csv.writer(si, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    if data: cw.writerow(data[0].keys())
    else: cw.writerow(['Projekt', 'Kupac', 'Sklop', 'GreskaID', 'TipGreske', 'OpisGreske', 'Evidentirao', 'VrijemeEvidentiranja', 'Rijesio', 'VrijemeRjesenja', 'SlikeOriginalniNazivi'])
    for row in data:
        row_list = list(row)
        if row_list[7]: dt_obj = format_datetime_hr_excel(row_list[7]); row_list[7] = dt_obj.strftime('%d.%m.%Y. %H:%M:%S') if isinstance(dt_obj, datetime.datetime) else row_list[7]
        if row_list[9]: dt_obj = format_datetime_hr_excel(row_list[9]); row_list[9] = dt_obj.strftime('%d.%m.%Y. %H:%M:%S') if isinstance(dt_obj, datetime.datetime) else row_list[9]
        cw.writerow(row_list)
    output = si.getvalue(); si.close(); timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S"); filename = f"evidencija_gresaka_{timestamp}.csv"
    return Response( output.encode('utf-8-sig'), mimetype='text/csv', headers={ "Content-Disposition": f"attachment;filename={filename}", "Content-Type": "text/csv; charset=utf-8-sig" })

@app.route('/api/stats/errors_by_customer_and_type')
@login_required
def error_stats_by_customer_and_type():
    db = get_db()
    query = """ SELECT g.tip AS TipGreske, p.kupac AS Kupac FROM greske g JOIN sklopovi s ON g.sklop_id = s.id JOIN projekti p ON s.projekt_id = p.id WHERE g.tip IS NOT NULL AND g.tip != '' AND p.kupac IS NOT NULL AND p.kupac != ''; """
    cursor = db.execute(query); results = cursor.fetchall(); stats = {}
    for row in results:
        kupac = row['Kupac']; tip = row['TipGreske']
        if kupac not in stats: stats[kupac] = Counter()
        stats[kupac][tip] += 1
    final_stats = {kupac: dict(tipovi) for kupac, tipovi in stats.items()}; return jsonify(final_stats)

# === 7. Pokretanje Aplikacije ===
if __name__ == '__main__':
    print("Provjera i inicijalizacija baze prije pokretanja...")
    with app.app_context():
        init_db(get_db())
    print("Pokretanje Flask servera...")

    app.run(debug=True, host='0.0.0.0', port=5000)
